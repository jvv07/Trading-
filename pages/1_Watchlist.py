import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import yfinance as yf

from lib.supabase_client import get_client, SOLO_USER_ID
from lib.indicators import sma, ema, rsi, macd, bollinger_bands
from lib.charts import candlestick_with_indicators
from lib.style import (
    inject_css, kpi_card, section_header, info_banner, stat_row,
    company_card_header, score_bar, check_item, analyst_badge, valuation_model_card,
)
from lib.nav import render_nav
from lib.fundamental import (
    fetch_info, fetch_financials, fetch_holders, fetch_market_data,
    fetch_peer_info, fetch_fmp, get_fmp_key, safe_get, format_large,
    bs_row, _first_val,
    calc_dcf, calc_graham_number, calc_ddm, calc_altman_z, calc_relative_valuation,
    score_value, score_future, score_past, score_health, score_dividend,
    get_sector_peers, SECTOR_NAME_MAP,
)

st.set_page_config(page_title="Watchlist / Research", layout="wide",
                   initial_sidebar_state="expanded")
inject_css()
render_nav("Watchlist")
st.title("Watchlist & Research")

client = get_client()

# ── Watchlist CRUD ────────────────────────────────────────────────────────────

def load_watchlist():
    res = (client.table("watchlist").select("*")
           .eq("user_id", SOLO_USER_ID).order("added_at").execute())
    return res.data or []


with st.sidebar:
    st.subheader("Add Symbol")
    with st.form("add_watch"):
        new_sym   = st.text_input("Ticker").upper().strip()
        new_notes = st.text_input("Notes (optional)")
        add_btn   = st.form_submit_button("Add", use_container_width=True)
    if add_btn and new_sym:
        try:
            client.table("watchlist").insert({
                "user_id": SOLO_USER_ID,
                "symbol":  new_sym,
                "notes":   new_notes or None,
            }).execute()
            st.rerun()
        except Exception:
            st.warning(f"{new_sym} already in watchlist.")

    st.divider()
    st.caption("Chart overlay (used in candlestick view)")
    chart_period = st.selectbox("Period", ["1mo","3mo","6mo","1y","2y","5y"], index=3)
    show_sma20   = st.checkbox("SMA 20",  value=True)
    show_sma50   = st.checkbox("SMA 50",  value=True)
    show_sma200  = st.checkbox("SMA 200", value=False)
    show_ema     = st.checkbox("EMA 20",  value=False)
    show_bb      = st.checkbox("Bollinger Bands", value=False)
    show_rsi     = st.checkbox("RSI (14)", value=True)
    show_macd    = st.checkbox("MACD",    value=False)

watchlist = load_watchlist()

# ── Global search: ?ticker=X routing ──────────────────────────────────────────
_search_ticker = st.query_params.get("ticker", "").upper().strip()
if _search_ticker:
    # Direct research view for searched ticker — bypass watchlist selection
    _in_wl = any(w["symbol"] == _search_ticker for w in watchlist)
    _add_col, _back_col, _ = st.columns([2, 2, 8])
    with _back_col:
        if st.button("← Back to Watchlist", type="secondary"):
            st.query_params.clear()
            st.rerun()
    with _add_col:
        if not _in_wl:
            if st.button(f"+ Add {_search_ticker} to Watchlist", type="primary"):
                try:
                    client.table("watchlist").insert({
                        "user_id": SOLO_USER_ID, "symbol": _search_ticker,
                    }).execute()
                    st.success(f"{_search_ticker} added to watchlist.")
                except Exception:
                    st.warning(f"{_search_ticker} already in watchlist.")
        else:
            st.success(f"✓ {_search_ticker} is in your watchlist")
    st.divider()
    selected_sym = _search_ticker
    wl_entry = None
else:
    if not watchlist:
        st.info("Add symbols using the sidebar.")
        st.stop()

# ── Snapshot expander ─────────────────────────────────────────────────────────

@st.cache_data(ttl=300)
def _snapshot(symbols_key: str):
    symbols = symbols_key.split(",")
    rows = []
    for sym in symbols:
        try:
            fi   = yf.Ticker(sym).fast_info
            last = fi.last_price
            prev = fi.previous_close
            chg  = last - prev
            pct  = chg / prev * 100 if prev else 0
            rows.append({"Symbol": sym, "Last": last, "Chg": chg, "Chg%": pct,
                         "52W Hi": fi.year_high, "52W Lo": fi.year_low,
                         "vs Hi%": (last / fi.year_high - 1)*100 if fi.year_high else None})
        except Exception:
            rows.append({"Symbol": sym, "Last": None, "Chg": None, "Chg%": None,
                         "52W Hi": None, "52W Lo": None, "vs Hi%": None})
    return rows


@st.cache_data(ttl=3600)
def _rsi_snapshot(syms_key: str) -> dict:
    result = {}
    for sym in syms_key.split(","):
        try:
            df = yf.download(sym, period="3mo", auto_adjust=True, progress=False)
            if isinstance(df.columns, pd.MultiIndex):
                df.columns = df.columns.droplevel(1)
            r = rsi(df["Close"])
            result[sym] = round(float(r.iloc[-1]), 1) if not r.empty else None
        except Exception:
            result[sym] = None
    return result


if not _search_ticker:
 syms_key = ",".join(sorted(w["symbol"] for w in watchlist))

if not _search_ticker and watchlist:
 with st.expander("Watchlist Snapshot", expanded=False):
    snap     = _snapshot(syms_key)
    snap_df  = pd.DataFrame(snap)
    rsi_vals = _rsi_snapshot(syms_key)
    snap_df["RSI"] = snap_df["Symbol"].map(rsi_vals)

    def _rsi_color(v):
        if v is None or pd.isna(v): return ""
        if v >= 70: return "color:#ff4b4b"
        if v <= 30: return "color:#00d4aa"
        return ""

    st.dataframe(
        snap_df.style
            .format({
                "Last":   lambda v: f"${v:.2f}" if v else "N/A",
                "Chg":    lambda v: f"{v:+.2f}" if v else "N/A",
                "Chg%":   lambda v: f"{v:+.2f}%" if v else "N/A",
                "52W Hi": lambda v: f"${v:.2f}" if v else "N/A",
                "52W Lo": lambda v: f"${v:.2f}" if v else "N/A",
                "vs Hi%": lambda v: f"{v:.1f}%" if v else "N/A",
                "RSI":    lambda v: f"{v:.1f}" if v else "N/A",
            })
            .map(lambda v: f"color:{'#00d4aa' if isinstance(v,str) and v.startswith('+') else '#ff4b4b' if isinstance(v,str) and v.startswith('-') and v != '-' else ''}",
                 subset=["Chg","Chg%"])
            .map(_rsi_color, subset=["RSI"]),
        use_container_width=True, hide_index=True,
    )

if not _search_ticker:
    st.divider()
    # ── Symbol selector ──────────────────────────────────────────────────────
    sym_list     = [w["symbol"] for w in watchlist]
    selected_sym = st.radio(
        "Select symbol for research report:",
        sym_list,
        horizontal=True,
        label_visibility="visible",
    )
    wl_entry = next((w for w in watchlist if w["symbol"] == selected_sym), None)
    col_remove, _ = st.columns([1, 6])
    if col_remove.button(f"Remove {selected_sym}", type="secondary"):
        client.table("watchlist").delete().eq("user_id", SOLO_USER_ID).eq("symbol", selected_sym).execute()
        st.rerun()
    st.divider()

# ── Load all data ─────────────────────────────────────────────────────────────

# Safe defaults — used if data loading partially fails
_empty_fin = {k: pd.DataFrame() for k in
              ["annual_income","quarterly_income","annual_bs",
               "quarterly_bs","annual_cf","quarterly_cf"]}
info        = {}
financials  = _empty_fin.copy()
holders     = {"institutional": pd.DataFrame(), "major": pd.DataFrame(),
               "insider_tx": pd.DataFrame()}
market_d    = {"recommendations": pd.DataFrame(), "upgrades": pd.DataFrame(),
               "news": [], "dividends": pd.Series(dtype=float)}
fmp_data    = {}
_load_error = None

with st.spinner(f"Loading {selected_sym} report…"):
    try:
        info       = fetch_info(selected_sym)
        financials = fetch_financials(selected_sym)
        holders    = fetch_holders(selected_sym)
        market_d   = fetch_market_data(selected_sym)
        fmp_data   = fetch_fmp(selected_sym)
    except Exception as _e:
        _load_error = str(_e)

# Run scoring (safe — each returns a default tuple on error)
def _safe_score(fn, *args):
    try:
        return fn(*args)
    except Exception:
        return 0.0, []

val_score,  val_sigs  = _safe_score(score_value,    info, financials)
fut_score,  fut_sigs  = _safe_score(score_future,   info, financials, fmp_data)
past_score, past_sigs = _safe_score(score_past,     info, financials)
hlt_score,  hlt_sigs  = _safe_score(score_health,   info, financials)
div_score,  div_sigs  = _safe_score(score_dividend, info, financials)

try:
    dcf_result = calc_dcf(info, financials)
except Exception:
    dcf_result = None
try:
    graham = calc_graham_number(info)
except Exception:
    graham = None
try:
    ddm_result = calc_ddm(info)
except Exception:
    ddm_result = None
try:
    altman = calc_altman_z(info, financials)
except Exception:
    altman = None
try:
    rel_val = calc_relative_valuation(info, financials)
except Exception:
    rel_val = {"pe_fair_value": None, "evebitda_fair_value": None,
               "pe_median": 20, "ev_median": 14, "sector": ""}
peers = get_sector_peers(safe_get(info, "sector", ""), selected_sym)

if _load_error:
    st.info("Some data could not be fetched — partial results may be shown.")

# ── Company Research Report ────────────────────────────────────────────────────

st.html(
    "<div style='color:#00d4aa;font-size:.7rem;font-weight:700;"
    "text-transform:uppercase;letter-spacing:.15em;margin-bottom:.5rem'>"
    f"◆ COMPANY RESEARCH REPORT — {selected_sym}</div>")

left_col, right_col = st.columns([1, 2.5], gap="large")

# ════════════════════════════════════════════════════════════════════════════════
#  LEFT COLUMN
# ════════════════════════════════════════════════════════════════════════════════
with left_col:
    # Company card
    price      = safe_get(info, "currentPrice") or safe_get(info, "regularMarketPrice")
    prev_close = safe_get(info, "previousClose") or safe_get(info, "regularMarketPreviousClose")
    chg_pct    = ((float(price) - float(prev_close)) / float(prev_close) * 100
                  if price and prev_close else None)
    mkt_cap    = safe_get(info, "marketCap")

    st.html(company_card_header(
        ticker       = selected_sym,
        name         = safe_get(info, "longName", safe_get(info, "shortName", selected_sym)),
        sector       = safe_get(info, "sector", "N/A"),
        industry     = safe_get(info, "industry", "N/A"),
        employees    = safe_get(info, "fullTimeEmployees"),
        market_cap_str = format_large(mkt_cap),
        price        = price,
        change_pct   = chg_pct,
    ))

    # Pentagon radar
    dims       = ["Value", "Future", "Past", "Health", "Dividend"]
    dim_scores = [val_score, fut_score, past_score, hlt_score, div_score]
    r_scores   = dim_scores + [dim_scores[0]]  # close polygon
    theta      = dims + [dims[0]]

    fig_radar = go.Figure()
    # Background
    fig_radar.add_trace(go.Scatterpolar(
        r=[6, 6, 6, 6, 6, 6], theta=theta,
        fill="toself", fillcolor="rgba(26,35,50,0.5)",
        line=dict(color="#1a2332", width=1),
        showlegend=False, hoverinfo="skip",
    ))
    # Score polygon
    fig_radar.add_trace(go.Scatterpolar(
        r=r_scores, theta=theta,
        fill="toself", fillcolor="rgba(0,212,170,0.15)",
        line=dict(color="#00d4aa", width=2),
        showlegend=False,
        hovertemplate="%{theta}: %{r:.1f}/6<extra></extra>",
    ))
    fig_radar.update_layout(
        polar=dict(
            bgcolor="rgba(0,0,0,0)",
            radialaxis=dict(
                range=[0, 6], dtick=2, showticklabels=True,
                tickfont=dict(size=9, color="#4a5a72"),
                gridcolor="#1a2332", linecolor="#1a2332",
            ),
            angularaxis=dict(
                tickfont=dict(size=11, color="#8892a4"),
                gridcolor="#1a2332", linecolor="#1a2332",
            ),
        ),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font_color="#e2e8f0",
        margin=dict(l=30, r=30, t=20, b=20),
        height=280,
        showlegend=False,
    )
    st.plotly_chart(fig_radar, use_container_width=True)

    # Score bars
    st.markdown("---")
    for label, sc in zip(dims, dim_scores):
        st.html(score_bar(label, sc))

    # Overall score
    total = sum(dim_scores)
    overall = (total / 30) * 10
    ov_color = "#ff4b4b" if overall < 3.5 else "#f1c14e" if overall < 6 else "#00d4aa"
    st.html(f"""
<div style="background:linear-gradient(135deg,#0d1422,#0a1020);
            border:1px solid #1a2332;border-radius:12px;padding:14px;
            text-align:center;margin-top:12px">
  <div style="color:#4a5a72;font-size:.68rem;font-weight:700;
              text-transform:uppercase;letter-spacing:.1em">Overall Score</div>
  <div style="color:{ov_color};font-size:2.2rem;font-weight:900;
              letter-spacing:-.04em">{overall:.1f}<span style="font-size:1rem;
              color:#4a5a72">/10</span></div>
</div>""")

    # Price target bar
    t_lo  = safe_get(info, "targetLowPrice")
    t_hi  = safe_get(info, "targetHighPrice")
    t_med = safe_get(info, "targetMeanPrice")
    if t_lo and t_hi and price:
        st.markdown("---")
        st.html("<div style='color:#4a5a72;font-size:.68rem;font-weight:700;"
                    "text-transform:uppercase;letter-spacing:.1em;margin-bottom:8px'>"
                    "Analyst Price Targets</div>")
        fig_pt = go.Figure()
        fig_pt.add_trace(go.Bar(
            x=["Low","Mean","High","Current"],
            y=[t_lo, t_med or 0, t_hi, float(price)],
            marker_color=["#4a5a72","#f1c14e","#00d4aa","#4e9af1"],
            text=[f"${v:.0f}" for v in [t_lo, t_med or 0, t_hi, float(price)]],
            textposition="outside", textfont=dict(size=10),
        ))
        fig_pt.update_layout(
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            font_color="#e2e8f0", showlegend=False,
            margin=dict(l=0, r=0, t=10, b=20),
            height=180,
            yaxis=dict(gridcolor="#1a2332", showticklabels=False),
            xaxis=dict(gridcolor="rgba(0,0,0,0)"),
        )
        st.plotly_chart(fig_pt, use_container_width=True)

    if wl_entry and wl_entry.get("notes"):
        st.html(info_banner(f"📌 {wl_entry['notes']}", "#4e9af1"))


# ════════════════════════════════════════════════════════════════════════════════
#  RIGHT COLUMN — 8 TABS
# ════════════════════════════════════════════════════════════════════════════════
with right_col:
    tabs = st.tabs([
        "Overview", "Valuation", "Future Growth", "Past Performance",
        "Financial Health", "Dividend", "Management", "Ownership",
    "Consensus"])

    # ── TAB 0: OVERVIEW ───────────────────────────────────────────────────────
    with tabs[0]:
        try:
            desc = safe_get(info, "longBusinessSummary", "No description available.")
            with st.expander("Business Description", expanded=True):
                st.write(desc)

            # KPI row
            pe_str    = f"{safe_get(info,'trailingPE'):.1f}x" if safe_get(info,"trailingPE") else "N/A"
            eps_str   = f"${safe_get(info,'trailingEps'):.2f}" if safe_get(info,"trailingEps") else "N/A"
            beta_str  = f"{safe_get(info,'beta'):.2f}" if safe_get(info,"beta") else "N/A"
            hi52      = safe_get(info, "fiftyTwoWeekHigh")
            lo52      = safe_get(info, "fiftyTwoWeekLow")
            rng52     = (f"${lo52:.2f} – ${hi52:.2f}" if hi52 and lo52 else "N/A")
            vol_str   = format_large(safe_get(info,"volume"), prefix="")
            rec_key   = safe_get(info, "recommendationKey", "")
            n_analysts = safe_get(info, "numberOfAnalystOpinions", 0) or 0

            st.html(stat_row([
                ("P/E (TTM)", pe_str),
                ("EPS (TTM)", eps_str),
                ("Beta",      beta_str),
                ("52W Range", rng52),
                ("Volume",    vol_str),
            ]))

            c1, c2 = st.columns([1, 2])
            with c1:
                st.html(
                    analyst_badge(rec_key) +
                    f'<span style="color:#4a5a72;font-size:.75rem;margin-left:8px">'
                    f'{n_analysts} analysts</span>')
            st.markdown("")

            # Peer comparison table
            if peers:
                st.html(section_header("Peer Comparison"))
                with st.spinner("Loading peers…"):
                    p_info = fetch_peer_info(",".join(sorted(peers)))
                peer_rows = []
                all_tickers = [selected_sym] + list(p_info.keys())
                all_infos   = {selected_sym: info, **p_info}
                for sym2 in all_tickers:
                    inf2  = all_infos[sym2]
                    peer_rows.append({
                        "Symbol":   sym2,
                        "Price":    safe_get(inf2,"currentPrice") or safe_get(inf2,"regularMarketPrice"),
                        "Mkt Cap":  format_large(safe_get(inf2,"marketCap")),
                        "P/E":      f"{safe_get(inf2,'trailingPE'):.1f}" if safe_get(inf2,"trailingPE") else "N/A",
                        "Fwd P/E":  f"{safe_get(inf2,'forwardPE'):.1f}"  if safe_get(inf2,"forwardPE")  else "N/A",
                        "Rev Gr%":  f"{safe_get(inf2,'revenueGrowth',0)*100:.1f}%" if safe_get(inf2,"revenueGrowth") else "N/A",
                        "Net Mgn%": f"{safe_get(inf2,'profitMargins',0)*100:.1f}%" if safe_get(inf2,"profitMargins") else "N/A",
                    })
                peer_df = pd.DataFrame(peer_rows)
                st.dataframe(peer_df, use_container_width=True, hide_index=True)

            # News feed
            news = market_d.get("news", [])
            if news:
                st.html(section_header("Latest News"))
                for item in news[:6]:
                    try:
                        title = item.get("title","")
                        link  = item.get("link","")
                        pub   = item.get("publisher","")
                        st.html(
                            f'<div style="padding:6px 0;border-bottom:1px solid #1a2332">'
                            f'<a href="{link}" target="_blank" style="color:#e2e8f0;'
                            f'text-decoration:none;font-size:.85rem;font-weight:500">{title}</a>'
                            f'<div style="color:#4a5a72;font-size:.7rem;margin-top:2px">{pub}</div>'
                            f'</div>')
                    except Exception:
                        pass

        except Exception as e:
            st.info("Overview data could not be loaded for this symbol.")

    # ── TAB 1: VALUATION ──────────────────────────────────────────────────────
    with tabs[1]:
        try:
            price = safe_get(info, "currentPrice") or safe_get(info, "regularMarketPrice")

            def _upside(fv):
                if fv is None or price is None: return None
                return (float(fv) - float(price)) / float(price) * 100

            # 4 model cards
            c1, c2 = st.columns(2, gap="medium")
            with c1:
                st.html(valuation_model_card(
                    "DCF (2-Stage)",
                    dcf_result["fair_value"] if dcf_result else None,
                    price,
                    _upside(dcf_result["fair_value"] if dcf_result else None),
                    dcf_result["methodology"] if dcf_result else "Requires positive FCF",
                ))
            with c2:
                st.html(valuation_model_card(
                    "Graham Number",
                    graham["graham_number"] if graham else None,
                    price,
                    _upside(graham["graham_number"] if graham else None),
                    "√(22.5 × EPS × Book Value)",
                ))

            st.markdown("")
            c3, c4 = st.columns(2, gap="medium")
            rel = calc_relative_valuation(info, financials)
            with c3:
                st.html(valuation_model_card(
                    f"Relative P/E (vs sector {rel['pe_median']}×)",
                    rel["pe_fair_value"],
                    price,
                    _upside(rel["pe_fair_value"]),
                    f"Sector median P/E: {rel['pe_median']}×",
                ))
            with c4:
                st.html(valuation_model_card(
                    f"EV/EBITDA (vs {rel['ev_median']}×)",
                    rel["evebitda_fair_value"],
                    price,
                    _upside(rel["evebitda_fair_value"]),
                    f"Sector median EV/EBITDA: {rel['ev_median']}×",
                ))

            if ddm_result:
                st.markdown("")
                st.html(valuation_model_card(
                    "DDM (Dividend Discount)",
                    ddm_result["fair_value"],
                    price,
                    _upside(ddm_result["fair_value"]),
                    ddm_result["methodology"],
                    accent="#f1c14e",
                ))

            # Price vs fair values bar chart
            st.markdown("---")
            fv_vals, fv_labels, fv_colors = [], [], []
            if price:
                fv_vals.append(float(price)); fv_labels.append("Current"); fv_colors.append("#4e9af1")
            if dcf_result:
                fv_vals.append(dcf_result["fair_value"]); fv_labels.append("DCF"); fv_colors.append("#00d4aa")
            if graham:
                fv_vals.append(graham["graham_number"]); fv_labels.append("Graham"); fv_colors.append("#00b894")
            if rel["pe_fair_value"]:
                fv_vals.append(rel["pe_fair_value"]); fv_labels.append("Rel P/E"); fv_colors.append("#f1c14e")
            if rel["evebitda_fair_value"]:
                fv_vals.append(rel["evebitda_fair_value"]); fv_labels.append("EV/EBITDA"); fv_colors.append("#a29bfe")
            if ddm_result:
                fv_vals.append(ddm_result["fair_value"]); fv_labels.append("DDM"); fv_colors.append("#fd79a8")

            if len(fv_vals) > 1:
                fig_fv = go.Figure(go.Bar(
                    x=fv_labels, y=fv_vals,
                    marker_color=fv_colors,
                    text=[f"${v:.2f}" for v in fv_vals],
                    textposition="outside",
                ))
                fig_fv.update_layout(
                    title="Price vs Valuation Models",
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    font_color="#e2e8f0", showlegend=False,
                    yaxis=dict(gridcolor="#1a2332", tickprefix="$"),
                    xaxis=dict(gridcolor="rgba(0,0,0,0)"),
                    margin=dict(l=0, r=0, t=40, b=20), height=300,
                )
                st.plotly_chart(fig_fv, use_container_width=True)

            # Valuation signals
            st.html(section_header("Valuation Signals"))
            for sig in val_sigs:
                st.html(check_item(sig["text"], sig["passed"]))

            # Peer scatter: P/E vs Revenue Growth
            if peers:
                st.html(section_header("Peer: P/E vs Revenue Growth"))
                p_info = fetch_peer_info(",".join(sorted(peers)))
                all_info2 = {selected_sym: info, **p_info}
                sx, sy, slabels = [], [], []
                for sym2, inf2 in all_info2.items():
                    pe2 = safe_get(inf2, "trailingPE")
                    gr2 = safe_get(inf2, "revenueGrowth")
                    if pe2 and gr2:
                        sx.append(pe2); sy.append(gr2*100); slabels.append(sym2)
                if sx:
                    fig_sc = go.Figure(go.Scatter(
                        x=sx, y=sy, mode="markers+text",
                        text=slabels, textposition="top center",
                        marker=dict(
                            color=["#00d4aa" if s == selected_sym else "#4e9af1" for s in slabels],
                            size=10,
                        ),
                        textfont=dict(size=10, color="#8892a4"),
                    ))
                    fig_sc.update_layout(
                        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                        font_color="#e2e8f0",
                        xaxis=dict(title="P/E Ratio", gridcolor="#1a2332"),
                        yaxis=dict(title="Revenue Growth %", gridcolor="#1a2332"),
                        margin=dict(l=0, r=0, t=20, b=20), height=300,
                    )
                    st.plotly_chart(fig_sc, use_container_width=True)

        except Exception as e:
            st.info("Valuation data could not be loaded for this symbol.")

    # ── TAB 2: FUTURE GROWTH ──────────────────────────────────────────────────
    with tabs[2]:
        try:
            rev_g    = safe_get(info, "revenueGrowth")
            earn_g   = safe_get(info, "earningsGrowth")
            fwd_eps  = safe_get(info, "forwardEps")
            trail_eps = safe_get(info, "trailingEps")

            # Fallback: calculate YoY growth from income statement when info dict lacks them
            _inc_fg = financials.get("annual_income", pd.DataFrame())
            if rev_g is None:
                _rr = bs_row(_inc_fg, "Total Revenue", "Revenue")
                if _rr is not None:
                    _rv = _rr.dropna()
                    if len(_rv) >= 2 and float(_rv.iloc[1]) != 0:
                        rev_g = (float(_rv.iloc[0]) - float(_rv.iloc[1])) / abs(float(_rv.iloc[1]))
            if earn_g is None:
                _nr = bs_row(_inc_fg, "Net Income")
                if _nr is not None:
                    _nv = _nr.dropna()
                    if len(_nv) >= 2 and float(_nv.iloc[1]) > 0:
                        earn_g = (float(_nv.iloc[0]) - float(_nv.iloc[1])) / abs(float(_nv.iloc[1]))

            st.html(stat_row([
                ("Revenue Growth", f"{rev_g*100:.1f}%"   if rev_g is not None else "N/A",
                 "#00d4aa" if rev_g and rev_g > 0 else "#ff4b4b"),
                ("Earnings Growth", f"{earn_g*100:.1f}%" if earn_g is not None else "N/A",
                 "#00d4aa" if earn_g and earn_g > 0 else "#ff4b4b"),
                ("Fwd EPS",  f"${fwd_eps:.2f}"  if fwd_eps  else "N/A"),
                ("Trail EPS", f"${trail_eps:.2f}" if trail_eps else "N/A"),
                ("Target",   f"${safe_get(info,'targetMeanPrice'):.2f}"
                              if safe_get(info,"targetMeanPrice") else "N/A", "#f1c14e"),
            ]))

            # Revenue + Earnings bar chart (4Y)
            inc = financials.get("annual_income", pd.DataFrame())
            rev_row = bs_row(inc, "Total Revenue", "Revenue")
            ni_row  = bs_row(inc, "Net Income")

            if rev_row is not None and not rev_row.dropna().empty:
                rev_data = rev_row.dropna().iloc[:4]
                ni_data  = ni_row.dropna().iloc[:4]  if ni_row is not None else pd.Series(dtype=float)

                years = [str(d)[:4] for d in rev_data.index][::-1]
                revs  = [float(v)/1e9 for v in rev_data.values][::-1]
                nis   = [float(v)/1e9 for v in ni_data.reindex(rev_data.index).values][::-1] if not ni_data.empty else []

                fig_gr = make_subplots(specs=[[{"secondary_y": True}]])
                fig_gr.add_trace(go.Bar(name="Revenue ($B)", x=years, y=revs,
                                        marker_color="#4e9af1"), secondary_y=False)
                if nis:
                    ni_colors = ["#00d4aa" if v >= 0 else "#ff4b4b" for v in nis]
                    fig_gr.add_trace(go.Bar(name="Net Income ($B)", x=years, y=nis,
                                            marker_color=ni_colors), secondary_y=True)

                # FMP forecast bars
                fmp_ests = fmp_data.get("analyst_estimates", [])
                if fmp_ests and get_fmp_key():
                    try:
                        fwd_revs = [float(e.get("estimatedRevenueAvg", 0))/1e9 for e in fmp_ests[:2]]
                        fwd_yrs  = [e.get("date","")[:4] for e in fmp_ests[:2]]
                        fig_gr.add_trace(go.Bar(name="Est Revenue ($B)", x=fwd_yrs, y=fwd_revs,
                                                marker_color="#00d4aa", opacity=0.5),
                                         secondary_y=False)
                    except Exception:
                        pass

                fig_gr.update_layout(
                    title="Annual Revenue & Net Income", barmode="group",
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    font_color="#e2e8f0", legend=dict(bgcolor="rgba(0,0,0,0)"),
                    yaxis=dict(title="Revenue $B", gridcolor="#1a2332"),
                    yaxis2=dict(title="Net Income $B", gridcolor="rgba(0,0,0,0)"),
                    margin=dict(l=0, r=0, t=40, b=20), height=320,
                )
                st.plotly_chart(fig_gr, use_container_width=True)
            else:
                st.info("Annual income statement not available.")

            # Upgrades/downgrades
            upgrades = market_d.get("upgrades", pd.DataFrame())
            if not upgrades.empty:
                st.html(section_header("Recent Upgrades / Downgrades"))
                try:
                    disp = upgrades.reset_index()
                    st.dataframe(disp.head(10), use_container_width=True, hide_index=True)
                except Exception:
                    st.dataframe(upgrades.head(10), use_container_width=True)

            # Future signals
            st.html(section_header("Future Growth Signals"))
            for sig in fut_sigs:
                st.html(check_item(sig["text"], sig["passed"]))

        except Exception as e:
            st.info("Growth data could not be loaded for this symbol.")

    # ── TAB 3: PAST PERFORMANCE ───────────────────────────────────────────────
    with tabs[3]:
        try:
            inc = financials.get("annual_income", pd.DataFrame())
            cf  = financials.get("annual_cf", pd.DataFrame())

            roe = safe_get(info, "returnOnEquity")
            roa = safe_get(info, "returnOnAssets")
            net_mgn  = safe_get(info, "profitMargins")
            gross_mgn = safe_get(info, "grossMargins")
            op_mgn   = safe_get(info, "operatingMargins")
            # Fallback: calculate margins from income statement
            if any(v is None for v in [net_mgn, gross_mgn, op_mgn]):
                _rev_r = bs_row(inc, "Total Revenue", "Revenue")
                if _rev_r is not None and not _rev_r.dropna().empty:
                    _rev_v = float(_rev_r.dropna().iloc[0])
                    if _rev_v != 0:
                        if net_mgn is None:
                            _ni = _first_val(bs_row(inc, "Net Income"))
                            if _ni is not None: net_mgn = _ni / _rev_v
                        if gross_mgn is None:
                            _gp = _first_val(bs_row(inc, "Gross Profit"))
                            if _gp is not None: gross_mgn = _gp / _rev_v
                        if op_mgn is None:
                            _oi = _first_val(bs_row(inc, "Operating Income", "Operating Income Loss"))
                            if _oi is not None: op_mgn = _oi / _rev_v
            st.html(stat_row([
                ("ROE",        f"{roe*100:.1f}%" if roe is not None else "N/A",
                 "#00d4aa" if roe and roe > 0.15 else "#f1c14e" if roe and roe > 0 else "#ff4b4b"),
                ("ROA",        f"{roa*100:.1f}%" if roa is not None else "N/A",
                 "#00d4aa" if roa and roa > 0.05 else "#f1c14e" if roa and roa > 0 else "#ff4b4b"),
                ("Net Margin", f"{net_mgn*100:.1f}%"   if net_mgn  is not None else "N/A"),
                ("Gross Mgn",  f"{gross_mgn*100:.1f}%" if gross_mgn is not None else "N/A"),
                ("Op Margin",  f"{op_mgn*100:.1f}%"    if op_mgn   is not None else "N/A"),
            ]))

            # Revenue + NI grouped bars
            rev_row = bs_row(inc, "Total Revenue", "Revenue")
            ni_row  = bs_row(inc, "Net Income")
            if rev_row is not None and not rev_row.dropna().empty:
                rev_data = rev_row.dropna().iloc[:5]
                years = [str(d)[:4] for d in rev_data.index][::-1]
                revs  = [float(v)/1e9 for v in rev_data.values][::-1]

                fig_hist = go.Figure()
                fig_hist.add_trace(go.Bar(name="Revenue ($B)", x=years, y=revs,
                                           marker_color="#4e9af1"))
                if ni_row is not None and not ni_row.dropna().empty:
                    ni_data = ni_row.dropna().reindex(rev_data.index)
                    nis = [float(v)/1e9 if not pd.isna(v) else 0 for v in ni_data.values][::-1]
                    ni_colors = ["#00d4aa" if v >= 0 else "#ff4b4b" for v in nis]
                    fig_hist.add_trace(go.Bar(name="Net Income ($B)", x=years, y=nis,
                                              marker_color=ni_colors))
                fig_hist.update_layout(
                    title="Revenue & Net Income History", barmode="group",
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    font_color="#e2e8f0", legend=dict(bgcolor="rgba(0,0,0,0)"),
                    yaxis=dict(gridcolor="#1a2332"),
                    margin=dict(l=0, r=0, t=40, b=20), height=300,
                )
                st.plotly_chart(fig_hist, use_container_width=True)

            # Margin trend
            gross_row = bs_row(inc, "Gross Profit")
            op_row    = bs_row(inc, "Operating Income", "Operating Income Loss")
            ni_row2   = bs_row(inc, "Net Income")
            if rev_row is not None and gross_row is not None:
                rev_d = rev_row.dropna().iloc[:5]
                years = [str(d)[:4] for d in rev_d.index][::-1]
                fig_mg = go.Figure()
                for row_data, name, color in [
                    (gross_row, "Gross Margin", "#00d4aa"),
                    (op_row,    "Op Margin",    "#f1c14e"),
                    (ni_row2,   "Net Margin",   "#4e9af1"),
                ]:
                    if row_data is not None:
                        mgns = [(float(r)/float(v)*100) if float(v) != 0 else 0
                                for r, v in zip(
                                    row_data.reindex(rev_d.index).fillna(0).values,
                                    rev_d.values,
                                )][::-1]
                        fig_mg.add_trace(go.Scatter(
                            x=years, y=mgns, name=name, mode="lines+markers",
                            line=dict(color=color, width=2),
                        ))
                fig_mg.update_layout(
                    title="Margin Trends (%)",
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    font_color="#e2e8f0", legend=dict(bgcolor="rgba(0,0,0,0)"),
                    yaxis=dict(gridcolor="#1a2332", ticksuffix="%"),
                    xaxis=dict(gridcolor="#1a2332"),
                    margin=dict(l=0, r=0, t=40, b=20), height=280,
                )
                st.plotly_chart(fig_mg, use_container_width=True)

            # Altman Z components
            if altman:
                st.html(section_header(
                    f"Altman Z-Score: {altman['z_score']:.2f} — {altman['zone']}"))
                comps = altman["components"]
                fig_z = go.Figure(go.Bar(
                    x=list(comps.keys()),
                    y=list(comps.values()),
                    marker_color=["#00d4aa" if v > 0 else "#ff4b4b" for v in comps.values()],
                ))
                fig_z.update_layout(
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    font_color="#e2e8f0",
                    yaxis=dict(gridcolor="#1a2332"),
                    margin=dict(l=0, r=0, t=10, b=20), height=200,
                )
                st.plotly_chart(fig_z, use_container_width=True)

            # Past signals
            st.html(section_header("Past Performance Signals"))
            for sig in past_sigs:
                st.html(check_item(sig["text"], sig["passed"]))

        except Exception as e:
            st.info("Historical data could not be loaded for this symbol.")

    # ── TAB 4: FINANCIAL HEALTH ───────────────────────────────────────────────
    with tabs[4]:
        try:
            total_cash = float(safe_get(info, "totalCash", 0) or 0)
            total_debt = float(safe_get(info, "totalDebt", 0) or 0)
            curr_ratio = safe_get(info, "currentRatio")
            de         = safe_get(info, "debtToEquity")
            de_actual  = (de / 100) if de else None

            st.html(stat_row([
                ("Cash",          format_large(total_cash)),
                ("Total Debt",    format_large(total_debt)),
                ("Net Cash",      format_large(total_cash - total_debt),
                 "#00d4aa" if total_cash > total_debt else "#ff4b4b"),
                ("Current Ratio", f"{curr_ratio:.2f}" if curr_ratio else "N/A",
                 "#00d4aa" if curr_ratio and curr_ratio > 1.5 else
                 "#f1c14e" if curr_ratio and curr_ratio > 1 else "#ff4b4b"),
                ("D/E",           f"{de_actual:.2f}x" if de_actual else "N/A",
                 "#00d4aa" if de_actual and de_actual < 1 else
                 "#f1c14e" if de_actual and de_actual < 2 else "#ff4b4b"),
            ]))

            # Cash vs Debt bar
            fig_cd = go.Figure(go.Bar(
                x=["Cash", "Total Debt"],
                y=[total_cash/1e9, total_debt/1e9],
                marker_color=["#00d4aa", "#ff4b4b"],
                text=[f"${v/1e9:.2f}B" for v in [total_cash, total_debt]],
                textposition="outside",
            ))
            fig_cd.update_layout(
                title="Cash vs Total Debt ($B)",
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                font_color="#e2e8f0", showlegend=False,
                yaxis=dict(gridcolor="#1a2332", ticksuffix="B"),
                margin=dict(l=0, r=0, t=40, b=20), height=250,
            )
            st.plotly_chart(fig_cd, use_container_width=True)

            # Balance sheet composition
            bs = financials.get("annual_bs", pd.DataFrame())
            if not bs.empty:
                curr_assets = _first_val(bs_row(bs, "Current Assets", "Total Current Assets"))
                non_curr_a  = _first_val(bs_row(bs, "Net PPE", "Non Current Assets",
                                                 "Total Non Current Assets"))
                curr_liab   = _first_val(bs_row(bs, "Current Liabilities", "Total Current Liabilities"))
                non_curr_l  = _first_val(bs_row(bs, "Non Current Liabilities",
                                                  "Total Non Current Liabilities"))
                if curr_assets and curr_liab:
                    fig_bs = go.Figure()
                    fig_bs.add_trace(go.Bar(name="Current Assets",   x=["Assets"],
                                            y=[curr_assets/1e9],    marker_color="#00d4aa"))
                    fig_bs.add_trace(go.Bar(name="Non-Current Assets", x=["Assets"],
                                            y=[(non_curr_a or 0)/1e9], marker_color="#00a882"))
                    fig_bs.add_trace(go.Bar(name="Current Liab.",    x=["Liabilities"],
                                            y=[curr_liab/1e9],      marker_color="#ff4b4b"))
                    fig_bs.add_trace(go.Bar(name="Non-Current Liab.", x=["Liabilities"],
                                            y=[(non_curr_l or 0)/1e9], marker_color="#cc3333"))
                    fig_bs.update_layout(
                        title="Balance Sheet Composition ($B)", barmode="stack",
                        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                        font_color="#e2e8f0", legend=dict(bgcolor="rgba(0,0,0,0)"),
                        yaxis=dict(gridcolor="#1a2332", ticksuffix="B"),
                        margin=dict(l=0, r=0, t=40, b=20), height=280,
                    )
                    st.plotly_chart(fig_bs, use_container_width=True)

            # Altman Z gauge
            if altman:
                z = altman["z_score"]
                fig_gauge = go.Figure(go.Indicator(
                    mode="gauge+number",
                    value=min(z, 5),
                    title={"text": "Altman Z-Score", "font": {"color": "#8892a4", "size": 14}},
                    number={"font": {"color": altman["color"], "size": 28},
                            "suffix": f" ({altman['zone']})"},
                    gauge={
                        "axis": {"range": [0, 5], "tickcolor": "#4a5a72"},
                        "bar": {"color": altman["color"]},
                        "bgcolor": "#0d1422",
                        "steps": [
                            {"range": [0, 1.81], "color": "#2a0010"},
                            {"range": [1.81, 2.99], "color": "#2a2000"},
                            {"range": [2.99, 5], "color": "#003322"},
                        ],
                        "threshold": {"line": {"color": "#e2e8f0", "width": 2},
                                      "thickness": 0.75, "value": z},
                    },
                ))
                fig_gauge.update_layout(
                    paper_bgcolor="rgba(0,0,0,0)", font_color="#e2e8f0",
                    margin=dict(l=20, r=20, t=30, b=20), height=220,
                )
                st.plotly_chart(fig_gauge, use_container_width=True)

            # Health signals
            st.html(section_header("Financial Health Signals"))
            for sig in hlt_sigs:
                st.html(check_item(sig["text"], sig["passed"]))

        except Exception as e:
            st.info("Financial health data could not be loaded for this symbol.")

    # ── TAB 5: DIVIDEND ───────────────────────────────────────────────────────
    with tabs[5]:
        try:
            div_yield = safe_get(info, "dividendYield")
            div_rate  = safe_get(info, "dividendRate")

            if not div_yield or div_yield <= 0:
                st.html(info_banner(
                    "This company does not currently pay a dividend. "
                    "The dividend score is 0/6.", "#4a5a72"
                ))
            else:
                payout    = safe_get(info, "payoutRatio")
                ex_date   = safe_get(info, "exDividendDate")
                last_div  = safe_get(info, "lastDividendValue")
                five_yr   = safe_get(info, "fiveYearAvgDividendYield")

                st.html(stat_row([
                    ("Yield",        f"{div_yield*100:.2f}%", "#00d4aa"),
                    ("Annual Rate",  f"${div_rate:.2f}"       if div_rate  else "N/A"),
                    ("Payout Ratio", f"{payout*100:.0f}%"     if payout   else "N/A",
                     "#00d4aa" if payout and payout < 0.6 else
                     "#f1c14e" if payout and payout < 0.8 else "#ff4b4b"),
                    ("5Y Avg Yield", f"{five_yr:.2f}%"        if five_yr  else "N/A"),
                    ("Last Div",     f"${last_div:.4f}"       if last_div else "N/A"),
                ]))

                # Dividend history bar
                divs = market_d.get("dividends", pd.Series(dtype=float))
                if not divs.empty:
                    try:
                        annual = divs.resample("YE").sum()
                        annual = annual[annual > 0].tail(10)
                        years  = [str(d.year) for d in annual.index]
                        fig_div = go.Figure(go.Bar(
                            x=years, y=annual.values,
                            marker_color="#00d4aa",
                            text=[f"${v:.2f}" for v in annual.values],
                            textposition="outside",
                        ))
                        fig_div.add_trace(go.Scatter(
                            x=years, y=annual.values, mode="lines+markers",
                            line=dict(color="#f1c14e", width=2, dash="dot"),
                            name="Trend",
                        ))
                        fig_div.update_layout(
                            title="Annual Dividend per Share",
                            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                            font_color="#e2e8f0", showlegend=False,
                            yaxis=dict(gridcolor="#1a2332", tickprefix="$"),
                            margin=dict(l=0, r=0, t=40, b=20), height=280,
                        )
                        st.plotly_chart(fig_div, use_container_width=True)
                    except Exception:
                        pass

                # DDM sensitivity table (±1% growth / WACC grid)
                if ddm_result:
                    import numpy as np
                    st.html(section_header("DDM Sensitivity Table"))
                    base_g    = ddm_result["growth_rate"]
                    base_wacc = ddm_result["wacc"]
                    g_range   = [base_g - 0.01, base_g, base_g + 0.01]
                    w_range   = [base_wacc - 0.01, base_wacc, base_wacc + 0.01]
                    rows_ddm  = {}
                    for g in g_range:
                        row_data = {}
                        for w in w_range:
                            if w > g + 0.005 and div_rate:
                                fv = float(div_rate) * (1 + g) / (w - g)
                                row_data[f"WACC={w*100:.1f}%"] = f"${fv:.2f}"
                            else:
                                row_data[f"WACC={w*100:.1f}%"] = "N/A"
                        rows_ddm[f"g={g*100:.1f}%"] = row_data
                    st.dataframe(pd.DataFrame(rows_ddm).T, use_container_width=True)

            # Dividend signals
            st.html(section_header("Dividend Signals"))
            for sig in div_sigs:
                st.html(check_item(sig["text"], sig["passed"]))

        except Exception as e:
            st.info("Dividend data could not be loaded for this symbol.")

    # ── TAB 6: MANAGEMENT ─────────────────────────────────────────────────────
    with tabs[6]:
        try:
            # Governance risk scores
            audit_risk = safe_get(info, "auditRisk")
            board_risk = safe_get(info, "boardRisk")
            comp_risk  = safe_get(info, "compensationRisk")
            sh_rights  = safe_get(info, "shareHolderRightsRisk")
            overall_gr = safe_get(info, "overallRisk")

            gov_items = [
                ("Audit Risk",            audit_risk),
                ("Board Risk",            board_risk),
                ("Compensation Risk",     comp_risk),
                ("Shareholder Rights",    sh_rights),
                ("Overall Governance",    overall_gr),
            ]
            has_gov = any(v is not None for _, v in gov_items)
            if has_gov:
                st.html(section_header("Governance Risk (1=Low, 10=High)"))
                g_labels = [l for l, v in gov_items if v is not None]
                g_values = [float(v) for _, v in gov_items if v is not None]
                g_colors = ["#ff4b4b" if v >= 7 else "#f1c14e" if v >= 4 else "#00d4aa"
                            for v in g_values]
                fig_gov = go.Figure(go.Bar(
                    x=g_values, y=g_labels, orientation="h",
                    marker_color=g_colors,
                    text=[str(int(v)) for v in g_values], textposition="outside",
                ))
                fig_gov.update_layout(
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    font_color="#e2e8f0", showlegend=False,
                    xaxis=dict(range=[0, 11], gridcolor="#1a2332"),
                    yaxis=dict(gridcolor="rgba(0,0,0,0)"),
                    margin=dict(l=0, r=0, t=10, b=20), height=200,
                )
                st.plotly_chart(fig_gov, use_container_width=True)
            else:
                st.html(info_banner("Governance risk scores not available.", "#4a5a72"))

            # Insider transactions
            ins_tx = holders.get("insider_tx", pd.DataFrame())

            # Merge with FMP insider data if available
            fmp_ins = fmp_data.get("insider_trading", [])
            if fmp_ins:
                try:
                    fmp_df = pd.DataFrame(fmp_ins)[
                        ["reportingName", "transactionType", "securitiesTransacted",
                         "price", "transactionDate"]
                    ].rename(columns={
                        "reportingName":       "Name",
                        "transactionType":     "Transaction",
                        "securitiesTransacted":"Shares",
                        "price":               "Price",
                        "transactionDate":     "Date",
                    })
                    ins_tx = fmp_df
                except Exception:
                    pass

            if not ins_tx.empty:
                st.html(section_header("Insider Transactions"))
                st.dataframe(ins_tx.head(15), use_container_width=True, hide_index=True)

                # Buy/sell donut
                try:
                    tx_col = None
                    for col in ["Transaction", "Transaction Type", "Relationship"]:
                        if col in ins_tx.columns:
                            tx_col = col
                            break
                    if tx_col:
                        buys  = ins_tx[ins_tx[tx_col].str.contains("Buy|Purchase|S-1", case=False, na=False)].shape[0]
                        sells = ins_tx[ins_tx[tx_col].str.contains("Sell|Sale|S-1 K", case=False, na=False)].shape[0]
                        if buys + sells > 0:
                            fig_ins = go.Figure(go.Pie(
                                labels=["Buys", "Sells"],
                                values=[buys, sells],
                                marker_colors=["#00d4aa", "#ff4b4b"],
                                hole=0.5,
                            ))
                            fig_ins.update_layout(
                                paper_bgcolor="rgba(0,0,0,0)", font_color="#e2e8f0",
                                showlegend=True, legend=dict(bgcolor="rgba(0,0,0,0)"),
                                margin=dict(l=20, r=20, t=20, b=20), height=220,
                            )
                            st.plotly_chart(fig_ins, use_container_width=True)
                except Exception:
                    pass
            else:
                st.html(info_banner("Insider transaction data not available.", "#4a5a72"))

            # Analyst recommendation trend
            recs = market_d.get("recommendations", pd.DataFrame())
            if not recs.empty:
                st.html(section_header("Analyst Recommendation Trend"))
                try:
                    recs2 = recs.copy()
                    if "period" in recs2.columns:
                        recs2 = recs2.set_index("period")
                    grade_cols = [c for c in recs2.columns if c in
                                  ["strongBuy","buy","hold","sell","strongSell"]]
                    if grade_cols:
                        fig_rec = go.Figure()
                        colors  = {"strongBuy":"#00d4aa","buy":"#00b894","hold":"#f1c14e",
                                   "sell":"#e17055","strongSell":"#ff4b4b"}
                        for gc in grade_cols:
                            fig_rec.add_trace(go.Bar(
                                name=gc, x=recs2.index,
                                y=recs2[gc].values,
                                marker_color=colors.get(gc, "#4a5a72"),
                            ))
                        fig_rec.update_layout(
                            barmode="stack",
                            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                            font_color="#e2e8f0", legend=dict(bgcolor="rgba(0,0,0,0)"),
                            yaxis=dict(gridcolor="#1a2332"),
                            margin=dict(l=0, r=0, t=10, b=20), height=250,
                        )
                        st.plotly_chart(fig_rec, use_container_width=True)
                except Exception:
                    pass

        except Exception as e:
            st.info("Management data could not be loaded for this symbol.")

    # ── TAB 7: OWNERSHIP ──────────────────────────────────────────────────────
    with tabs[7]:
        try:
            inst = holders.get("institutional", pd.DataFrame())
            major = holders.get("major", pd.DataFrame())

            # Short interest
            short_pct = safe_get(info, "shortPercentOfFloat")
            short_str = f"{short_pct*100:.1f}%" if short_pct else "N/A"
            inst_pct  = safe_get(info, "heldPercentInstitutions")
            inside_pct = safe_get(info, "heldPercentInsiders")

            st.html(stat_row([
                ("Institutional %", f"{inst_pct*100:.1f}%"  if inst_pct   else "N/A"),
                ("Insider %",       f"{inside_pct*100:.1f}%" if inside_pct else "N/A"),
                ("Short Float %",   short_str,
                 "#ff4b4b" if short_pct and short_pct > 0.15 else
                 "#f1c14e" if short_pct and short_pct > 0.05 else "#00d4aa"),
            ]))

            # Major holders
            if not major.empty:
                st.html(section_header("Major Holders"))
                st.dataframe(major, use_container_width=True, hide_index=True)

            # Institutional holders
            if not inst.empty:
                col_a, col_b = st.columns([1.2, 1], gap="medium")
                with col_a:
                    st.html(section_header("Top Institutional Holders"))
                    disp_cols = [c for c in ["Holder", "Shares", "% Out", "Value"]
                                 if c in inst.columns]
                    if disp_cols:
                        st.dataframe(inst[disp_cols].head(15), use_container_width=True,
                                     hide_index=True)
                    else:
                        st.dataframe(inst.head(15), use_container_width=True, hide_index=True)

                with col_b:
                    # Top 10 horizontal bar
                    try:
                        bar_df = inst.head(10)
                        h_col  = next((c for c in ["Holder","Name"] if c in bar_df.columns), None)
                        s_col  = next((c for c in ["Shares","% Out"] if c in bar_df.columns), None)
                        if h_col and s_col:
                            fig_inst = go.Figure(go.Bar(
                                y=bar_df[h_col].str[:25],
                                x=bar_df[s_col],
                                orientation="h",
                                marker_color="#4e9af1",
                            ))
                            fig_inst.update_layout(
                                title="Top 10 Institutions",
                                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                                font_color="#e2e8f0", showlegend=False,
                                yaxis=dict(autorange="reversed", gridcolor="rgba(0,0,0,0)"),
                                xaxis=dict(gridcolor="#1a2332"),
                                margin=dict(l=0, r=0, t=40, b=20), height=350,
                            )
                            st.plotly_chart(fig_inst, use_container_width=True)
                    except Exception:
                        pass

                # Pie: top 5 institutional + "Other"
                try:
                    s_col2 = next((c for c in ["Shares","Value"] if c in inst.columns), None)
                    h_col2 = next((c for c in ["Holder","Name"] if c in inst.columns), None)
                    if s_col2 and h_col2:
                        top5   = inst.head(5)
                        others = inst.iloc[5:][s_col2].sum() if len(inst) > 5 else 0
                        pie_labels = list(top5[h_col2].str[:20]) + (["Others"] if others else [])
                        pie_vals   = list(top5[s_col2]) + ([others] if others else [])
                        fig_pie = go.Figure(go.Pie(
                            labels=pie_labels, values=pie_vals, hole=0.45,
                            textfont=dict(size=10),
                        ))
                        fig_pie.update_layout(
                            title="Institutional Ownership Distribution",
                            paper_bgcolor="rgba(0,0,0,0)", font_color="#e2e8f0",
                            legend=dict(bgcolor="rgba(0,0,0,0)", font=dict(size=10)),
                            margin=dict(l=0, r=0, t=40, b=20), height=300,
                        )
                        st.plotly_chart(fig_pie, use_container_width=True)
                except Exception:
                    pass
            else:
                st.html(info_banner("Institutional holder data not available.", "#4a5a72"))

        except Exception as e:
            st.info("Ownership data could not be loaded for this symbol.")

# ── Candlestick (legacy) in expander ─────────────────────────────────────────

with st.expander("Candlestick Chart", expanded=False):
    @st.cache_data(ttl=3600)
    def _chart_data(sym, period):
        df = yf.download(sym, period=period, auto_adjust=True, progress=False)
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.droplevel(1)
        return df

    df = _chart_data(selected_sym, chart_period)
    if df.empty:
        st.error(f"No data for {selected_sym}")
    else:
        price_indicators = {}
        if show_sma20:  price_indicators["SMA 20"]  = sma(df["Close"], 20)
        if show_sma50:  price_indicators["SMA 50"]  = sma(df["Close"], 50)
        if show_sma200: price_indicators["SMA 200"] = sma(df["Close"], 200)
        if show_ema:    price_indicators["EMA 20"]  = ema(df["Close"], 20)
        if show_bb:
            upper, middle, lower = bollinger_bands(df["Close"])
            price_indicators.update({"BB Upper": upper, "BB Middle": middle, "BB Lower": lower})
        rsi_indicators = {"RSI 14": rsi(df["Close"])} if show_rsi else {}
        all_indicators = {**price_indicators, **rsi_indicators}
        fig = candlestick_with_indicators(df, all_indicators,
                                          title=f"{selected_sym} — {chart_period}")
        st.plotly_chart(fig, use_container_width=True)

        if show_macd:
            macd_line, signal_line, histogram = macd(df["Close"])
            fig_macd = make_subplots(rows=1, cols=1)
            colors = ["#00d4aa" if v >= 0 else "#ff4b4b" for v in histogram]
            fig_macd.add_trace(go.Bar(x=df.index, y=histogram, name="Histogram",
                                      marker_color=colors))
            fig_macd.add_trace(go.Scatter(x=df.index, y=macd_line, name="MACD",
                                          line=dict(color="#4e9af1")))
            fig_macd.add_trace(go.Scatter(x=df.index, y=signal_line, name="Signal",
                                          line=dict(color="#f1c14e")))
            fig_macd.update_layout(
                title=f"{selected_sym} MACD",
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                font_color="#fafafa",
                xaxis=dict(gridcolor="#2a2f3e"), yaxis=dict(gridcolor="#2a2f3e"),
                margin=dict(l=10, r=10, t=40, b=10),
            )
            st.plotly_chart(fig_macd, use_container_width=True)


# =============================================================================
# Consensus Tab (auto-added by run_fix3.py)
# =============================================================================
try:
    import pandas as _cpd
    import io as _cio
    import plotly.graph_objects as _cgo
    from lib.fundamental import fetch_analyst_data as _fad, format_large as _fl, safe_get as _sg

    with tabs[8]:
        st.markdown("### Analyst Consensus")
        try:
            _ad = _fad(selected_sym)
        except Exception:
            _ad = {}

        _pr  = _sg(info,"currentPrice") or _sg(info,"regularMarketPrice") or 0
        _lo  = _sg(info,"targetLowPrice")
        _hi  = _sg(info,"targetHighPrice")
        _avg = _sg(info,"targetMeanPrice")
        _med = _sg(info,"targetMedianPrice")
        _nan = _sg(info,"numberOfAnalystOpinions")
        _rec = (_sg(info,"recommendationKey") or "").replace("_"," ").title()

        # KPI metrics
        _c1,_c2,_c3,_c4 = st.columns(4)
        _c1.metric("Target Low",  f"${float(_lo):.2f}"  if _lo  else "N/A")
        _c2.metric("Target Mean", f"${float(_avg):.2f}" if _avg else "N/A")
        _c3.metric("Target High", f"${float(_hi):.2f}"  if _hi  else "N/A")
        _up = ((float(_avg)-float(_pr))/float(_pr)*100) if (_avg and _pr) else None
        _c4.metric("Implied Upside", f"{_up:.1f}%" if _up is not None else "N/A",
                   delta=f"{_up:.1f}%" if _up is not None else None)
        st.caption(f"**Consensus:** {_rec or 'N/A'}  |  **Analysts:** {int(float(_nan)) if _nan else 'N/A'}")

        # Target range chart
        if _lo and _hi and _avg and _pr:
            _fig = _cgo.Figure()
            for _xv, _lb, _cl in [(float(_lo),"Low","#ff4b4b"),
                                   (float(_avg),"Mean","#00d4aa"),
                                   (float(_hi),"High","#4bffb5")]:
                _fig.add_vline(x=_xv, line_color=_cl, line_width=2,
                               annotation_text=f"{_lb} ${_xv:.0f}",
                               annotation_font_color=_cl)
            _fig.add_vrect(x0=float(_lo), x1=float(_hi),
                           fillcolor="rgba(0,212,170,0.08)", line_width=0)
            _fig.add_vline(x=float(_pr), line_dash="dash", line_color="#f1c14e",
                           line_width=2, annotation_text=f"Now ${float(_pr):.0f}",
                           annotation_font_color="#f1c14e")
            _fig.update_layout(height=90, showlegend=False,
                               margin=dict(t=30,b=5,l=10,r=10),
                               yaxis=dict(visible=False), xaxis=dict(tickprefix="$"),
                               paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)")
            st.plotly_chart(_fig, use_container_width=True)

        st.divider()

        # EPS estimates
        _eps_t = _ad.get("eps_trend", _cpd.DataFrame())
        if not _eps_t.empty:
            st.markdown("**EPS Estimates**")
            try: st.dataframe(_eps_t.style.format("{:.4f}", na_rep="—"), use_container_width=True)
            except Exception: st.dataframe(_eps_t, use_container_width=True)

        # Revenue estimates
        _rev_t = _ad.get("revenue_trend", _cpd.DataFrame())
        if not _rev_t.empty:
            st.markdown("**Revenue Estimates**")
            try: st.dataframe(_rev_t.style.format("{:,.0f}", na_rep="—"), use_container_width=True)
            except Exception: st.dataframe(_rev_t, use_container_width=True)

        # FMP estimates
        _fmp_e = _ad.get("fmp_estimates", [])
        if _fmp_e:
            st.markdown("**Analyst Estimates (FMP)**")
            _rows = [{"Date": _e.get("date",""),
                      "EPS Low": _e.get("estimatedEpsLow"),
                      "EPS Avg": _e.get("estimatedEpsAvg"),
                      "EPS High": _e.get("estimatedEpsHigh"),
                      "Rev Avg($B)": round(float(_e["estimatedRevenueAvg"])/1e9,2)
                                    if _e.get("estimatedRevenueAvg") else None}
                     for _e in _fmp_e[:8]]
            if _rows: st.dataframe(_cpd.DataFrame(_rows), use_container_width=True)

        # FMP price targets
        _fmp_t = _ad.get("fmp_targets", [])
        if _fmp_t:
            st.markdown("**Analyst Price Targets (FMP)**")
            _tr = [{"Date":_t.get("publishedDate","")[:10], "Analyst":_t.get("analystName",""),
                    "Company":_t.get("analystCompany",""), "Target":_t.get("priceTarget"),
                    "Action":_t.get("priceTargetDiff","")} for _t in _fmp_t[:15]]
            if _tr: st.dataframe(_cpd.DataFrame(_tr), use_container_width=True)

        # Rec trend
        _rec_t = _ad.get("rec_trend", _cpd.DataFrame())
        if not _rec_t.empty:
            st.divider(); st.markdown("**Recommendation Trend**")
            st.dataframe(_rec_t, use_container_width=True)

        # Upgrades/Downgrades
        try:
            _upgr = market_d.get("upgrades", _cpd.DataFrame())
            if not _upgr.empty:
                st.divider(); st.markdown("**Recent Upgrades/Downgrades**")
                _uc = [c for c in ["Firm","ToGrade","FromGrade","Action"] if c in _upgr.columns]
                if _uc: st.dataframe(_upgr[_uc].head(25), use_container_width=True)
        except Exception: pass

        # Financial Statements viewer
        st.divider()
        with st.expander("View Full Financial Statements"):
            try:
                _stmts = st.radio("Statement", ["Income Statement","Quarterly Income",
                    "Balance Sheet","Quarterly Balance Sheet","Cash Flow","Quarterly Cash Flow"],
                    horizontal=True)
                _dfmap = {"Income Statement": financials.get("annual_income",_cpd.DataFrame()),
                          "Quarterly Income":  financials.get("quarterly_income",_cpd.DataFrame()),
                          "Balance Sheet":     financials.get("annual_bs",_cpd.DataFrame()),
                          "Quarterly Balance Sheet": financials.get("quarterly_bs",_cpd.DataFrame()),
                          "Cash Flow":         financials.get("annual_cf",_cpd.DataFrame()),
                          "Quarterly Cash Flow": financials.get("quarterly_cf",_cpd.DataFrame())}
                _dfs = _dfmap.get(_stmts, _cpd.DataFrame())
                if not _dfs.empty:
                    try:
                        st.dataframe(_dfs.style.format(
                            lambda v: _fl(v,"") if isinstance(v,(int,float)) else str(v),
                            na_rep="—"), use_container_width=True)
                    except Exception:
                        st.dataframe(_dfs, use_container_width=True)
                else:
                    st.info("No data available for this statement.")
            except Exception: st.info("Financial statements unavailable.")

        # Excel downloads
        st.divider()
        _dl1, _dl2, _dl3 = st.columns(3)

        with _dl1:
            if st.button("Download Analyst Data", key="dl_cons"):
                try:
                    from openpyxl import Workbook as _WB
                    _wb = _WB(); _ws = _wb.active; _ws.title = "Price Targets"
                    for _r in [("Metric","Value"),("Symbol",selected_sym),
                               ("Price",float(_pr)),("Target Low",float(_lo) if _lo else None),
                               ("Target Mean",float(_avg) if _avg else None),
                               ("Target High",float(_hi) if _hi else None),
                               ("Upside %",round(_up,2) if _up else None),
                               ("Analysts",_nan),("Consensus",_rec)]:
                        _ws.append(_r)
                    if not _eps_t.empty:
                        _ws2 = _wb.create_sheet("EPS Estimates")
                        _ws2.append(["Period"]+list(_eps_t.columns))
                        for _i,_row in _eps_t.iterrows(): _ws2.append([str(_i)]+list(_row.values))
                    if _fmp_e:
                        _ws3 = _wb.create_sheet("FMP Estimates")
                        _ws3.append(["Date","EPS Low","EPS Avg","EPS High","Rev Avg"])
                        for _e in _fmp_e[:10]:
                            _ws3.append([_e.get("date"),_e.get("estimatedEpsLow"),
                                         _e.get("estimatedEpsAvg"),_e.get("estimatedEpsHigh"),
                                         _e.get("estimatedRevenueAvg")])
                    _buf = _cio.BytesIO(); _wb.save(_buf); _buf.seek(0)
                    st.download_button("Save",_buf.getvalue(),
                        f"{selected_sym}_consensus.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_cons_save")
                except Exception as _de: st.warning(f"Excel error: {_de}")

        with _dl2:
            if st.button("Download Financial Stmts", key="dl_fs"):
                try:
                    from openpyxl import Workbook as _WB2
                    _wb2 = _WB2()
                    _sheet_pairs = [
                        ("Income Stmt",  financials.get("annual_income",_cpd.DataFrame())),
                        ("Qtr Income",   financials.get("quarterly_income",_cpd.DataFrame())),
                        ("Balance Sheet",financials.get("annual_bs",_cpd.DataFrame())),
                        ("Qtr BS",       financials.get("quarterly_bs",_cpd.DataFrame())),
                        ("Cash Flow",    financials.get("annual_cf",_cpd.DataFrame())),
                        ("Qtr CF",       financials.get("quarterly_cf",_cpd.DataFrame())),
                    ]
                    _first = True
                    for _sn, _df in _sheet_pairs:
                        if _df is not None and not _df.empty:
                            _ws_f = _wb2.active if _first else _wb2.create_sheet(_sn)
                            if _first: _ws_f.title = _sn; _first = False
                            _ws_f.append(["Row"]+[str(c) for c in _df.columns])
                            for _ix,_rw in _df.iterrows():
                                _ws_f.append([str(_ix)]+[
                                    float(v) if isinstance(v,(int,float)) and v==v else None
                                    for v in _rw.values])
                    _buf2 = _cio.BytesIO(); _wb2.save(_buf2); _buf2.seek(0)
                    st.download_button("Save",_buf2.getvalue(),
                        f"{selected_sym}_financials.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_fs_save")
                except Exception as _fe2: st.warning(f"Excel error: {_fe2}")

        with _dl3:
            if st.button("Download DCF Model", key="dl_dcf"):
                try:
                    from openpyxl import Workbook as _WBd
                    from openpyxl.styles import Font as _Fnt, PatternFill as _Fill
                    from lib.fundamental import (calc_wacc as _cw, calc_dcf as _cd2,
                                                  calc_graham_number as _cgn,
                                                  bs_row as _bsr, _first_val as _fv)
                    _wb3 = _WBd()
                    # Sheet 1: Assumptions
                    _wa = _wb3.active; _wa.title = "Assumptions"
                    _wa.column_dimensions["A"].width = 28; _wa.column_dimensions["B"].width = 18
                    _hf = _Fnt(bold=True,color="FFFFFF"); _hfill = _Fill("solid",fgColor="1A2332")
                    _wa.append(["Assumption","Value"]); _wa["A1"].font=_hf; _wa["B1"].font=_hf
                    _wa["A1"].fill=_hfill; _wa["B1"].fill=_hfill
                    _pr_v = float(_sg(info,"currentPrice") or _sg(info,"regularMarketPrice") or 0)
                    _sh_v = float(_sg(info,"sharesOutstanding") or 0)
                    _bt_v = float(_sg(info,"beta") or 1.0)
                    _rg_v = float(_sg(info,"revenueGrowth") or 0.05)
                    _wacc_v = _cw(info)
                    _td_v = float(_sg(info,"totalDebt") or 0)
                    _tc_v = float(_sg(info,"totalCash") or 0)
                    for _ar in [("Symbol",selected_sym),("Current Price",_pr_v),
                                ("Shares Outstanding",_sh_v),("Beta",_bt_v),
                                ("Stage 1 Growth (g1)",round(_rg_v,4)),
                                ("Stage 2 Growth (g2)",round(_rg_v*0.6,4)),
                                ("Terminal Growth",0.025),("WACC",round(_wacc_v,4)),
                                ("Tax Rate",0.21),("Risk-Free Rate",0.045),
                                ("Equity Risk Premium",0.055),("Total Debt",_td_v),
                                ("Total Cash",_tc_v)]:
                        _wa.append(_ar)
                    # Sheet 2: DCF (row-based, formulas in col B)
                    _wd = _wb3.create_sheet("DCF Model")
                    _wd.column_dimensions["A"].width = 28; _wd.column_dimensions["B"].width = 18
                    _wd.append(["Item","Value"]); _wd["A1"].font=_hf; _wd["B1"].font=_hf
                    _wd["A1"].fill=_hfill; _wd["B1"].fill=_hfill
                    # Get base FCF
                    _cf = financials.get("annual_cf",_cpd.DataFrame())
                    _ocf_r = _bsr(_cf,"Operating Cash Flow","Cash From Operations","Total Cash From Operating Activities")
                    _cap_r = _bsr(_cf,"Capital Expenditure","Capital Expenditures","Purchase Of Property Plant And Equipment")
                    _ocf_b = _fv(_ocf_r) or float(_sg(info,"operatingCashflow") or 0)
                    _cap_b = abs(_fv(_cap_r) or 0)
                    _fcf_b = max(0, _ocf_b - _cap_b)
                    _wd.append(["Base FCF", _fcf_b])               # row 2 → B2
                    _wd.append(["Stage 1 Growth", "=Assumptions!B6"])  # B3
                    _wd.append(["Stage 2 Growth", "=Assumptions!B7"])  # B4
                    _wd.append(["WACC",           "=Assumptions!B9"])  # B5
                    _wd.append(["Terminal Growth","=Assumptions!B8"])  # B6
                    _wd.append(["",""])                                 # B7 spacer
                    # FCF projections rows 8-17 (years 1-10)
                    for _yr in range(1,11):
                        if _yr <= 5:
                            _f = f"=B2*(1+B3)^{_yr}"
                        else:
                            _f = f"=B2*(1+B3)^5*(1+B4)^{_yr-5}"
                        _wd.append([f"FCF Year {_yr}", _f])         # rows 8-17
                    # PV rows 18-27
                    for _yr in range(1,11):
                        _wd.append([f"PV Year {_yr}", f"=B{7+_yr}/(1+B5)^{_yr}"])
                    # Summary rows 28+
                    _wd.append(["Sum PV FCFs",     "=SUM(B18:B27)"])  # B28
                    _wd.append(["Terminal FCF",    "=B17*(1+B6)/(B5-B6)"])  # B29
                    _wd.append(["PV Terminal",     "=B29/(1+B5)^10"])  # B30
                    _wd.append(["Enterprise Value","=B28+B30"])         # B31
                    _wd.append(["Add: Cash",       _tc_v])              # B32
                    _wd.append(["Less: Debt",      _td_v])              # B33
                    _wd.append(["Equity Value",    "=B31+B32-B33"])     # B34
                    _wd.append(["Shares",          _sh_v])              # B35
                    _wd.append(["DCF Fair Value",  "=B34/B35"])         # B36
                    _wd.append(["Current Price",   _pr_v])              # B37
                    _wd.append(["Upside/(Downside)","=(B36-B37)/B37"])  # B38
                    # Sheet 3: Peer Comparison
                    _wp3 = _wb3.create_sheet("Peer Comparison")
                    _wp3.column_dimensions["A"].width=20; _wp3.column_dimensions["B"].width=14
                    _wp3.column_dimensions["C"].width=16; _wp3.column_dimensions["D"].width=14
                    _wp3.append(["Metric","Value","Sector Median","vs Median"])
                    _wp3["A1"].font=_hf; _wp3["B1"].font=_hf
                    _wp3["A1"].fill=_hfill; _wp3["B1"].fill=_hfill
                    _wp3["C1"].font=_hf; _wp3["D1"].font=_hf
                    _wp3["C1"].fill=_hfill; _wp3["D1"].fill=_hfill
                    from lib.fundamental import SECTOR_NAME_MAP as _snm, SECTOR_PE_MEDIANS as _spm, SECTOR_EV_EBITDA_MEDIANS as _sem
                    _sec = _sg(info,"sector","")
                    _msec = _snm.get(_sec, _sec)
                    _pe_m = _spm.get(_msec, 20); _ev_m = _sem.get(_msec, 14)
                    for _mn, _mv, _med in [
                        ("P/E (TTM)",    _sg(info,"trailingPE"),      _pe_m),
                        ("P/E (Forward)",_sg(info,"forwardPE"),       _pe_m),
                        ("EV/EBITDA",    _sg(info,"enterpriseToEbitda"),_ev_m),
                        ("P/B",          _sg(info,"priceToBook"),     None),
                        ("PEG",          _sg(info,"trailingPegRatio"),None),
                    ]:
                        try:
                            _v = float(_mv) if _mv else None
                            _vs = round((_v-_med)/_med*100,1) if (_v and _med) else None
                            _wp3.append([_mn, _v, _med, _vs])
                        except Exception: _wp3.append([_mn, None, _med, None])
                    _buf3 = _cio.BytesIO(); _wb3.save(_buf3); _buf3.seek(0)
                    st.download_button("Save",_buf3.getvalue(),
                        f"{selected_sym}_dcf_model.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_dcf_save")
                except Exception as _de3: st.warning(f"DCF model error: {_de3}")

except Exception:
    pass
